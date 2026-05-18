from pathlib import Path

import openpyxl

from app.config import settings
from app.core.types import FILE_SLOTS


def slot_path(slot: str) -> Path:
    return settings.upload_dir / FILE_SLOTS[slot]


def file_status(slot: str) -> dict:
    p = slot_path(slot)
    if not p.exists():
        return {"exists": False, "size": 0, "mtime": None, "sheets": None, "filename": FILE_SLOTS[slot]}
    st = p.stat()
    sheets = None
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        sheets = wb.sheetnames
        wb.close()
    except Exception:
        pass
    return {
        "exists": True,
        "size": st.st_size,
        "mtime": st.st_mtime,
        "sheets": sheets,
        "filename": FILE_SLOTS[slot],
    }


def all_files_status() -> dict[str, dict]:
    return {slot: file_status(slot) for slot in FILE_SLOTS}


def cache_key() -> str:
    mtimes = []
    for slot in FILE_SLOTS:
        p = slot_path(slot)
        if p.exists():
            mtimes.append(str(p.stat().st_mtime))
    return "|".join(sorted(mtimes)) if mtimes else "empty"


def detect_dates_from_swift() -> list[str]:
    """Read sheet names from swift_di file to auto-detect available dates."""
    p = slot_path("swift_di")
    if not p.exists():
        return []
    try:
        wb = openpyxl.load_workbook(str(p), read_only=True, data_only=True)
        names = wb.sheetnames
        wb.close()
        # Keep only sheets that look like DD.MM (e.g. '01.02')
        return [s for s in names if len(s) == 5 and s[2] == "." and s[:2].isdigit() and s[3:].isdigit()]
    except Exception:
        return []
