"""
Build individual transaction rows from source Excel files.
Returns a list of row dicts matching the frontend data model.
"""
from __future__ import annotations
import re
import unicodedata
from pathlib import Path

import openpyxl

from app.config import settings
from app.core.types import FILE_SLOTS
from app.modules.ingestion.service import cache_key

CRED_PAT = re.compile(r'06800-(\d+)-(\d+)\s+TRANSFER\s+CREDMBNEO\.(\d+)\.(\d+)')
DEB_PAT  = re.compile(r'06800-(\d+)-(\d+)\s+TRANSFER\s+')

_cache_key: str | None = None
_cache_rows: list[dict] = []


def _to_int(v):
    if v is None:
        return None
    try:
        return int(float(str(v)))
    except Exception:
        return None


def _to_str(v) -> str:
    if v is None:
        return ''
    iv = _to_int(v)
    return str(iv) if iv is not None else str(v).strip()


def _dt8_fmt(raw) -> str | None:
    """YYYYMMDD int/str → 'DD/MM/YYYY'"""
    if raw is None:
        return None
    s = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
    if len(s) == 8:
        return f'{s[6:8]}/{s[4:6]}/{s[0:4]}'
    return None


def _dt_iso_fmt(s) -> str | None:
    """'YYYY-MM-DD ...' → 'DD/MM/YYYY'"""
    if not s:
        return None
    p = str(s).strip().split(' ')[0]
    if '-' in p:
        parts = p.split('-')
        return f'{parts[2][:2]}/{parts[1]}/{parts[0]}'
    return None


def _parse_swift_di_time(raw) -> tuple[str | None, str | None]:
    """'M/D/YYYY H:MM:SS AM/PM' → ('DD/MM/YYYY', 'HH:MM')"""
    try:
        s = str(raw).strip()
        date_part, time_part = s.split(' ', 1)
        m, d, y = date_part.split('/')
        tp = time_part.strip().split(' ')
        hms = tp[0].split(':')
        h, mn = int(hms[0]), int(hms[1])
        if len(tp) > 1:
            if tp[1] == 'PM' and h != 12:
                h += 12
            elif tp[1] == 'AM' and h == 12:
                h = 0
        return f'{int(d):02d}/{int(m):02d}/{y}', f'{h:02d}:{mn:02d}'
    except Exception:
        return None, None


def _parse_swift_den_time(raw) -> tuple[str | None, str | None]:
    """'YYYY-MM-DD HH:MM:SS' → ('DD/MM/YYYY', 'HH:MM')"""
    try:
        s = str(raw).strip()
        dp, tp = s.split(' ')
        y, mo, d = dp.split('-')
        h, mn = tp.split(':')[0], tp.split(':')[1]
        return f'{d}/{mo}/{y}', f'{int(h):02d}:{int(mn):02d}'
    except Exception:
        return None, None


def _napas_time_fmt(raw) -> str | None:
    """HHMMSS int → 'HH:MM'"""
    try:
        s = str(int(raw)).zfill(6)
        return f'{s[0:2]}:{s[2:4]}'
    except Exception:
        return None


def _napas_date(raw, p_mmdd: str) -> tuple[str | None, str]:
    """MMDD int + processing MMDD → ('DD/MM/YYYY', 'GD'|'QT')"""
    try:
        s = str(int(raw)).zfill(4)
        mm, dd = s[0:2], s[2:4]
        return f'{dd}/{mm}/2026', 'GD' if s[0:4] == p_mmdd else 'QT'
    except Exception:
        return None, 'GD'


def _sheet_to_day(sh: str) -> str:
    """'01.02' → '01/02/2026'"""
    parts = sh.split('.')
    return f'{parts[0]}/{parts[1]}/2026'


def _swift_status(st_raw: str) -> str:
    # Strip diacritics so "THÀNH CÔNG" and "THANH CONG" both match
    u = unicodedata.normalize('NFD', str(st_raw)).encode('ascii', 'ignore').decode().upper()
    if 'THANH' in u:
        return 'THANH_CONG'
    if 'TIMEOUT' in u:
        return 'TIMEOUT'
    return 'THAT_BAI'


def _infer_recon_status(swift_st, core_entry, napas_info, napas_failed, swift_date) -> str:
    if swift_st == 'THAT_BAI':
        # Swift AND (usually) NAPAS both say this failed — but if Core has a
        # matching entry anyway, money may have posted despite both failure
        # reports. That's a real anomaly, not a normal cancelled transaction.
        return 'SWIFT_THAT_BAI_CO_CORE' if core_entry else 'SWIFT_THAT_BAI'
    if swift_st == 'TIMEOUT' and not core_entry:
        return 'SWIFT_TIMEOUT'
    if napas_failed and swift_st == 'THANH_CONG':
        # NAPAS reports KTC (failed) — normally Core has no matching entry.
        # If Core DOES have one anyway, that's a contradiction worth its own
        # status: money may have posted despite NAPAS reporting failure.
        return 'NAPAS_THAT_BAI_CO_CORE' if core_entry else 'NAPAS_THAT_BAI'
    if swift_st == 'TIMEOUT' and core_entry:
        return 'TIMEOUT_CO_CORE'
    if not core_entry and not napas_info and swift_st == 'THANH_CONG':
        return 'CHI_SWIFT'
    if core_entry and napas_info:
        cd, nd = core_entry.get('date'), napas_info.get('date')
        same = (swift_date == cd) and (swift_date == nd or nd == cd)
        return 'KHOP' if same else 'KHOP_LECH_NGAY'
    return 'CHI_SWIFT'


def _active_sheets(wb, allowed_mmdd: set[str] | None) -> list[str]:
    """Return sheet names to process, filtered to allowed_mmdd set if provided."""
    if allowed_mmdd is None:
        return wb.sheetnames
    return [s for s in wb.sheetnames if s[:4] in allowed_mmdd]


def _build(year: int = 2026, date_labels: list[str] | None = None) -> list[dict]:
    base_dir: Path = settings.upload_dir

    # Derive MMDD keys from date_labels (e.g. '01.02' → '0102')
    allowed_mmdd: set[str] | None = None
    if date_labels:
        allowed_mmdd = set()
        for dl in date_labels:
            parts = dl.split('.')
            if len(parts) == 2:
                allowed_mmdd.add(parts[1] + parts[0])   # MMDD

    # ── NAPAS indexes (scoped to allowed dates) ───────────────────────────────
    napas_di: dict[str, dict] = {}
    napas_den: dict[str, dict] = {}
    napas_ktc: dict[str, dict] = {}

    p_ndi = base_dir / FILE_SLOTS['napas_di']
    if p_ndi.exists():
        wb = openpyxl.load_workbook(str(p_ndi), read_only=True, data_only=True)
        for sh in _active_sheets(wb, allowed_mmdd):
            p_mmdd = sh[:4]
            for row in wb[sh].iter_rows(min_row=2, values_only=True):
                trace = _to_str(row[2])
                if not trace:
                    continue
                amt = _to_int(row[1]) or 0
                date_str, typ = _napas_date(row[4], p_mmdd) if row[4] else (None, 'GD')
                ntime = _napas_time_fmt(row[3]) if row[3] else None
                napas_di[trace] = {'amount': amt, 'date': date_str, 'failed': False, 'type': typ, 'time': ntime}
        wb.close()

    p_nden = base_dir / FILE_SLOTS['napas_den']
    if p_nden.exists():
        wb = openpyxl.load_workbook(str(p_nden), read_only=True, data_only=True)
        for sh in _active_sheets(wb, allowed_mmdd):
            p_mmdd = sh[:4]
            for row in wb[sh].iter_rows(min_row=2, values_only=True):
                trace = _to_str(row[3])
                if not trace:
                    continue
                amt = _to_int(row[2]) or 0
                date_str, typ = _napas_date(row[5], p_mmdd) if row[5] else (None, 'GD')
                ntime = _napas_time_fmt(row[4]) if row[4] else None
                napas_den[trace] = {'amount': amt, 'date': date_str, 'failed': False, 'type': typ, 'time': ntime}
        wb.close()

    p_ktc = base_dir / FILE_SLOTS['napas_di_fail']
    if p_ktc.exists():
        wb = openpyxl.load_workbook(str(p_ktc), read_only=True, data_only=True)
        for sh in _active_sheets(wb, allowed_mmdd):
            p_mmdd = sh[:4]
            for row in wb[sh].iter_rows(min_row=2, values_only=True):
                trace = _to_str(row[3])
                if not trace:
                    continue
                amt = _to_int(row[2]) or 0
                date_str, _ = _napas_date(row[5], p_mmdd) if row[5] else (None, 'GD')
                ntime = _napas_time_fmt(row[4]) if row[4] else None
                napas_ktc[trace] = {'amount': amt, 'date': date_str, 'failed': True, 'type': 'GD', 'time': ntime}
        wb.close()

    # ── Core GL indexes ───────────────────────────────────────────────────────
    core_cred: dict[str, dict] = {}  # seq → {trace, amount, date}
    core_deb: dict[str, dict] = {}   # seq → {amount, date}

    p_core = base_dir / FILE_SLOTS['core']
    if p_core.exists():
        wb = openpyxl.load_workbook(str(p_core), read_only=True, data_only=True)
        for sh in wb.sheetnames:
            for row in wb[sh].iter_rows(min_row=6, values_only=True):
                if row[1] is None or row[7] is None:
                    continue
                debit  = float(row[4]) if row[4] else 0
                credit = float(row[5]) if row[5] else 0
                desc   = str(row[7])
                core_dt = _dt8_fmt(row[1])
                if not core_dt:
                    continue
                if credit > 0:
                    m = CRED_PAT.search(desc)
                    if m:
                        core_cred[m.group(2)] = {'trace': m.group(4), 'amount': int(credit), 'date': core_dt}
                elif debit > 0:
                    m = DEB_PAT.search(desc)
                    if m:
                        core_deb[m.group(2)] = {'amount': int(debit), 'date': core_dt}
        wb.close()

    # ── Build rows from Swift ─────────────────────────────────────────────────
    rows: list[dict] = []
    rid = 1

    # Swift sheets are named DD.MM (first 5 chars), e.g. '01.02'
    allowed_ddmm: set[str] | None = set(date_labels) if date_labels else None

    p_sdi = base_dir / FILE_SLOTS['swift_di']
    if p_sdi.exists():
        wb = openpyxl.load_workbook(str(p_sdi), read_only=True, data_only=True)
        for sh in wb.sheetnames:
            if allowed_ddmm and sh[:5] not in allowed_ddmm:
                continue
            day_label = _sheet_to_day(sh[:5])
            for row in wb[sh].iter_rows(min_row=8, values_only=True):
                txn_date, _ = _parse_swift_di_time(row[0]) if row[0] else (None, None)
                trace = _to_str(row[6])
                amt   = _to_int(row[7]) or 0
                seq   = _to_str(row[12])
                hd    = _to_str(row[14])
                st    = _swift_status(str(row[16]) if row[16] else '')
                if not trace or not amt:
                    continue
                swift_date = _dt8_fmt(hd) if hd and len(hd) == 8 else txn_date

                core_e = core_cred.get(seq) if seq else None
                if core_e and core_e['amount'] != amt:
                    core_e = None
                n_tc = napas_di.get(trace)
                if n_tc and n_tc['amount'] != amt:
                    n_tc = None
                is_ktc = trace in napas_ktc
                if st == 'THAT_BAI':
                    core_e = None
                    n_tc = None

                n_info = napas_ktc.get(trace) if is_ktc else n_tc
                rs = _infer_recon_status(st, core_e, n_info, is_ktc, swift_date)

                rows.append({
                    'id': f'r{rid:03d}',
                    'trace': trace.zfill(6),
                    'sequence': seq or None,
                    'direction': 'Đi',
                    'amount': amt,
                    'day': day_label,
                    'swift': {'date': swift_date, 'txnDate': txn_date, 'status': st},
                    'core': {'date': core_e['date'], 'entry': 'Ghi có'} if core_e else None,
                    'napas': {
                        'date': n_info['date'], 'time': n_info.get('time'),
                        'failed': n_info['failed'], 'type': n_info.get('type', 'GD'),
                    } if n_info else None,
                    'recon_status': rs,
                    'resolved_by': None, 'resolved_at': None, 'note': None,
                })
                rid += 1
        wb.close()

    p_sden = base_dir / FILE_SLOTS['swift_den']
    if p_sden.exists():
        wb = openpyxl.load_workbook(str(p_sden), read_only=True, data_only=True)
        for sh in wb.sheetnames:
            if allowed_ddmm and sh[:5] not in allowed_ddmm:
                continue
            day_label = _sheet_to_day(sh[:5])
            for row in wb[sh].iter_rows(min_row=8, values_only=True):
                txn_date, _ = _parse_swift_den_time(row[1]) if row[1] else (None, None)
                amt   = _to_int(row[8]) or 0
                hd    = str(row[10]) if row[10] else ''
                trace = _to_str(row[11])
                seq   = _to_str(row[13])
                st    = _swift_status(str(row[16]) if row[16] else '')
                if not trace or not amt:
                    continue
                swift_date = _dt_iso_fmt(hd) if hd else txn_date

                core_e = core_deb.get(seq) if seq else None
                if core_e and core_e['amount'] != amt:
                    core_e = None
                n_tc = napas_den.get(trace)
                if n_tc and n_tc['amount'] != amt:
                    n_tc = None
                if st == 'THAT_BAI':
                    core_e = None
                    n_tc = None

                rs = _infer_recon_status(st, core_e, n_tc, False, swift_date)

                rows.append({
                    'id': f'r{rid:03d}',
                    'trace': trace.zfill(6),
                    'sequence': seq or None,
                    'direction': 'Đến',
                    'amount': amt,
                    'day': day_label,
                    'swift': {'date': swift_date, 'txnDate': txn_date, 'status': st},
                    'core': {'date': core_e['date'], 'entry': 'Ghi nợ'} if core_e else None,
                    'napas': {
                        'date': n_tc['date'], 'time': n_tc.get('time'),
                        'failed': False, 'type': n_tc.get('type', 'GD'),
                    } if n_tc else None,
                    'recon_status': rs,
                    'resolved_by': None, 'resolved_at': None, 'note': None,
                })
                rid += 1
        wb.close()

    return rows


def get_rows(year: int = 2026, date_labels: list[str] | None = None) -> list[dict]:
    global _cache_key, _cache_rows
    if date_labels:
        return _build(year, date_labels)
    ck = cache_key()
    if ck != _cache_key:
        _cache_rows = _build(year)
        _cache_key = ck
    return _cache_rows


def clear_rows_cache() -> None:
    global _cache_key
    _cache_key = None
