"""
Extract 10 real transactions per day (30 total) from example Excel files.
Outputs a JS array for data/reconcile.js INITIAL_ROWS.
New fields: swift.txnDate (actual transaction date), napas.time, day (processing day)
"""
import openpyxl, sys, io, glob, os, re

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

allfiles = glob.glob(r'd:\FoxAI\FoxAI\*\vcbneo-app\example\*')
fmap = {}
for f in allfiles:
    enc = os.path.basename(f).encode('utf-8')
    if b'Swift report \xc4\x91i.xlsx' == enc: fmap['swift_di'] = f
    elif b'Swift report \xc4\x91\xe1\xba\xbfn.xlsx' == enc: fmap['swift_den'] = f
    elif b'Napas \xc4\x91i.xlsx' == enc: fmap['napas_di'] = f
    elif b'Napas \xc4\x91\xe1\xba\xbfn.xlsx' == enc: fmap['napas_den'] = f
    elif b'Core.xlsx' == enc: fmap['core'] = f
    elif b'Napas \xc4\x91i KTC 20260201.XLSX' == enc: fmap['napas_ktc'] = f

CRED_PAT = re.compile(r'06800-(\d+)-(\d+)\s+TRANSFER\s+CREDMBNEO\.(\d+)\.(\d+)')
DEB_PAT  = re.compile(r'06800-(\d+)-(\d+)\s+TRANSFER\s+')

def dt8(raw):
    s = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
    return f'{s[6:8]}/{s[4:6]}/{s[0:4]}' if len(s) == 8 else s

def dt_iso(s):
    p = str(s).strip()
    if '-' in p:
        pp = p.split('-')
        return f'{pp[2][:2]}/{pp[1]}/{pp[0]}'
    return p

def parse_txn_time_di(raw):
    """Swift Đi THỜI GIAN: 'M/D/YYYY H:MM:SS AM/PM' -> ('DD/MM/YYYY', 'HH:MM')"""
    try:
        s = str(raw).strip()
        # e.g. '2/1/2026 11:31:11 PM'
        date_part, time_part = s.split(' ', 1)
        m, d, y = date_part.split('/')
        txn_date = f'{int(d):02d}/{int(m):02d}/{y}'
        # time part: '11:31:11 PM' or '10:45:46 AM'
        tp = time_part.strip()
        parts = tp.split(' ')
        hms = parts[0].split(':')
        h, mn = int(hms[0]), int(hms[1])
        if len(parts) > 1:
            if parts[1] == 'PM' and h != 12: h += 12
            elif parts[1] == 'AM' and h == 12: h = 0
        txn_time = f'{h:02d}:{mn:02d}'
        return txn_date, txn_time
    except:
        return None, None

def parse_txn_time_den(raw):
    """Swift Đến THỜI GIAN: 'YYYY-MM-DD HH:MM:SS' -> ('DD/MM/YYYY', 'HH:MM')"""
    try:
        s = str(raw).strip()
        # e.g. '2026-02-03 23:55:12'
        date_p, time_p = s.split(' ')
        y, m, d = date_p.split('-')
        txn_date = f'{d}/{m}/{y}'
        h, mn = time_p.split(':')[0], time_p.split(':')[1]
        txn_time = f'{int(h):02d}:{int(mn):02d}'
        return txn_date, txn_time
    except:
        return None, None

def napas_time_fmt(raw):
    """HHMMSS -> 'HH:MM'"""
    try:
        s = str(int(raw)).zfill(6) if isinstance(raw, (int, float)) else str(raw).strip().zfill(6)
        return f'{s[0:2]}:{s[2:4]}'
    except:
        return None

def napas_date(raw, processing_day_mmdd):
    s = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
    s = s.zfill(4)
    mm, dd = s[0:2], s[2:4]
    date_str = f'{dd}/{mm}/2026'
    p_mm, p_dd = processing_day_mmdd[0:2], processing_day_mmdd[2:4]
    typ = 'GD' if (mm == p_mm and dd == p_dd) else 'QT'
    return date_str, typ

def to_int(v):
    if v is None: return None
    try: return int(float(str(v)))
    except: return None

def to_str(v):
    if v is None: return ''
    iv = to_int(v)
    return str(iv) if iv is not None else str(v).strip()

def sheet_to_day(sh):
    """'01.02' -> '01/02/2026'"""
    parts = sh.split('.')
    return f'{parts[0]}/{parts[1]}/2026'

# ── Build NAPAS indexes ───────────────────────────────────────────────────────
wb = openpyxl.load_workbook(fmap['napas_di'], read_only=True, data_only=True)
napas_di = {}
for sh in wb.sheetnames:
    p_mmdd = sh[:4]
    for row in wb[sh].iter_rows(min_row=2, values_only=True):
        trace = to_str(row[2])
        if not trace: continue
        amount = to_int(row[1]) or 0
        date_str, typ = napas_date(row[4], p_mmdd)
        ntime = napas_time_fmt(row[3]) if row[3] else None  # col[3] = time in NAPAS Đi? check below
        napas_di[trace] = {'amount': amount, 'date': date_str, 'failed': False, 'type': typ, 'time': ntime}

wb = openpyxl.load_workbook(fmap['napas_den'], read_only=True, data_only=True)
napas_den = {}
for sh in wb.sheetnames:
    p_mmdd = sh[:4]
    for row in wb[sh].iter_rows(min_row=2, values_only=True):
        trace = to_str(row[3])
        if not trace: continue
        amount = to_int(row[2]) or 0
        date_str, typ = napas_date(row[5], p_mmdd)
        ntime = napas_time_fmt(row[4]) if row[4] else None  # col[4] = time
        napas_den[trace] = {'amount': amount, 'date': date_str, 'failed': False, 'type': typ, 'time': ntime}

wb = openpyxl.load_workbook(fmap['napas_ktc'], read_only=True, data_only=True)
napas_ktc = {}
for sh in wb.sheetnames:
    p_mmdd = sh[:4]
    for row in wb[sh].iter_rows(min_row=2, values_only=True):
        trace = to_str(row[3])
        if not trace: continue
        amount = to_int(row[2]) or 0
        raw_date = row[5]
        date_str, _ = napas_date(raw_date, p_mmdd) if raw_date else (p_mmdd, 'GD')
        ntime = napas_time_fmt(row[4]) if row[4] else None
        napas_ktc[trace] = {'amount': amount, 'date': date_str, 'failed': True, 'type': 'GD', 'time': ntime}

# CORE GL
wb = openpyxl.load_workbook(fmap['core'], read_only=True, data_only=True)
core_cred = {}
core_deb  = {}
for sh in wb.sheetnames:
    for row in wb[sh].iter_rows(min_row=6, values_only=True):
        if row[1] is None or row[7] is None: continue
        debit  = float(row[4]) if row[4] else 0
        credit = float(row[5]) if row[5] else 0
        desc   = str(row[7])
        raw_dt = str(int(row[1])) if isinstance(row[1], (int, float)) else str(row[1])
        core_dt = dt8(raw_dt)
        if credit > 0:
            m = CRED_PAT.search(desc)
            if m:
                seq = m.group(2)
                core_cred[seq] = {'trace': m.group(4), 'amount': int(credit), 'date': core_dt}
        elif debit > 0:
            m = DEB_PAT.search(desc)
            if m:
                seq = m.group(2)
                core_deb[seq] = {'amount': int(debit), 'date': core_dt}

print(f'# Indexes: core_cred={len(core_cred)} core_deb={len(core_deb)}', file=sys.stderr)
print(f'# Indexes: napas_di={len(napas_di)} napas_den={len(napas_den)} napas_ktc={len(napas_ktc)}', file=sys.stderr)

def infer_recon_status(st, core_e, n_info, n_failed, swift_date):
    if st == 'THAT_BAI': return 'SWIFT_THAT_BAI'
    if st == 'TIMEOUT' and not core_e: return 'SWIFT_TIMEOUT'
    if n_failed and st == 'THANH_CONG': return 'NAPAS_THAT_BAI'
    if st == 'TIMEOUT' and core_e: return 'TIMEOUT_CO_CORE'
    if not core_e and not n_info and st == 'THANH_CONG': return 'CHI_SWIFT'
    if core_e and n_info:
        sd, cd, nd = swift_date, core_e['date'], n_info['date']
        same = (sd == cd) and (sd == nd or nd == cd)
        return 'KHOP' if same else 'KHOP_LECH_NGAY'
    return 'CHI_SWIFT'

results = []
rid = 1

# ── Swift DI ──────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(fmap['swift_di'], read_only=True, data_only=True)

for sh in wb.sheetnames[:3]:
    ws = wb[sh]
    day_label = sheet_to_day(sh)
    day_rows = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        txn_date, txn_time = parse_txn_time_di(row[0]) if row[0] else (None, None)
        trace = to_str(row[6]);  amt = to_int(row[7]) or 0
        seq   = to_str(row[12]); hd  = to_str(row[14])
        st_raw = str(row[16]) if row[16] else ''
        st = 'THANH_CONG' if 'THANH' in st_raw else ('TIMEOUT' if 'TIMEOUT' in st_raw else 'THAT_BAI')
        if not trace or not seq: continue
        hd_fmt = dt8(hd) if len(hd) == 8 else hd
        core_e = core_cred.get(seq)
        if core_e and core_e['amount'] != amt: core_e = None
        n_tc = napas_di.get(trace)
        if n_tc and n_tc['amount'] != amt: n_tc = None
        is_ktc = trace in napas_ktc
        if st == 'THAT_BAI': core_e = None; n_tc = None
        day_rows.append({'trace': trace, 'seq': seq, 'amount': amt,
                         'swift_date': hd_fmt, 'txn_date': txn_date, 'txn_time': txn_time,
                         'status': st, 'core': core_e, 'napas_tc': n_tc, 'is_ktc': is_ktc})

    full_ok  = [r for r in day_rows if r['core'] and r['napas_tc'] and r['status']=='THANH_CONG'
                and r['core']['date'] == r['swift_date'] and r['napas_tc']['date'] == r['swift_date']]
    lech_ok  = [r for r in day_rows if r['core'] and r['napas_tc'] and r['status']=='THANH_CONG'
                and not (r['core']['date'] == r['swift_date'] and r['napas_tc']['date'] == r['swift_date'])]
    ktc_rows = [r for r in day_rows if r['is_ktc'] and r['status']=='THANH_CONG']
    to_rows  = [r for r in day_rows if r['status']=='TIMEOUT' and r['core']]
    to_only  = [r for r in day_rows if r['status']=='TIMEOUT' and not r['core']]
    fail_rows= [r for r in day_rows if r['status']=='THAT_BAI']

    selected = (full_ok[:3] + lech_ok[:1] + ktc_rows[:1] + to_rows[:1] + to_only[:1] + fail_rows[:1])[:5]
    if len(selected) < 5:
        selected += [r for r in full_ok if r not in selected][:5-len(selected)]

    for r in selected:
        core_e = r['core'];  n_tc = r['napas_tc'];  is_ktc = r['is_ktc'];  st = r['status'];  sd = r['swift_date']
        n_info = napas_ktc.get(r['trace']) if is_ktc else n_tc
        n_failed = is_ktc
        rs = infer_recon_status(st, core_e, n_info, n_failed, sd)
        entry_c = {'date': core_e['date'], 'entry': 'Ghi co'} if core_e else None
        entry_n = None
        if n_info:
            entry_n = {'date': n_info['date'], 'time': n_info.get('time'), 'failed': n_info['failed'], 'type': n_info.get('type','GD')}
        results.append({'id': f'r{rid:03d}', 'trace': r['trace'].zfill(6), 'sequence': r['seq'],
                        'direction': 'Di', 'amount': r['amount'], 'day': day_label,
                        'swift': {'date': sd, 'txnDate': r['txn_date'], 'status': st},
                        'core': entry_c, 'napas': entry_n,
                        'recon_status': rs})
        rid += 1

# ── Swift DEN ─────────────────────────────────────────────────────────────────
wb = openpyxl.load_workbook(fmap['swift_den'], read_only=True, data_only=True)
for sh in wb.sheetnames[:3]:
    ws = wb[sh]
    day_label = sheet_to_day(sh)
    day_rows = []
    for row in ws.iter_rows(min_row=8, values_only=True):
        txn_date, txn_time = parse_txn_time_den(row[1]) if row[1] else (None, None)
        amt    = to_int(row[8]) or 0
        hd     = str(row[10]) if row[10] else ''
        trace  = to_str(row[11]); seq = to_str(row[13])
        st_raw = str(row[16]) if row[16] else ''
        st = 'THANH_CONG' if 'THANH' in st_raw else ('TIMEOUT' if 'TIMEOUT' in st_raw else 'THAT_BAI')
        if not trace or not seq: continue
        hd_fmt = dt_iso(hd)
        core_e = core_deb.get(seq)
        if core_e and core_e['amount'] != amt: core_e = None
        n_tc = napas_den.get(trace)
        if n_tc and n_tc['amount'] != amt: n_tc = None
        if st == 'THAT_BAI': core_e = None; n_tc = None
        day_rows.append({'trace': trace, 'seq': seq, 'amount': amt,
                         'swift_date': hd_fmt, 'txn_date': txn_date, 'txn_time': txn_time,
                         'status': st, 'core': core_e, 'napas_tc': n_tc})

    full_ok = [r for r in day_rows if r['core'] and r['napas_tc'] and r['status']=='THANH_CONG'
               and r['core']['date'] == r['swift_date'] and r['napas_tc']['date'] == r['swift_date']]
    lech_ok = [r for r in day_rows if r['core'] and r['napas_tc'] and r['status']=='THANH_CONG'
               and not (r['core']['date'] == r['swift_date'] and r['napas_tc']['date'] == r['swift_date'])]
    to_rows = [r for r in day_rows if r['status']=='TIMEOUT']
    chi_sw  = [r for r in day_rows if not r['core'] and not r['napas_tc'] and r['status']=='THANH_CONG']

    selected = (full_ok[:3] + lech_ok[:1] + to_rows[:1] + chi_sw[:1])[:5]
    if len(selected) < 5:
        selected += [r for r in full_ok if r not in selected][:5-len(selected)]
    if len(selected) < 5:
        selected += [r for r in chi_sw if r not in selected][:5-len(selected)]

    for r in selected:
        core_e = r['core'];  n_tc = r['napas_tc'];  st = r['status'];  sd = r['swift_date']
        rs = infer_recon_status(st, core_e, n_tc, False, sd)
        entry_c = {'date': core_e['date'], 'entry': 'Ghi no'} if core_e else None
        entry_n = None
        if n_tc:
            entry_n = {'date': n_tc['date'], 'time': n_tc.get('time'), 'failed': False, 'type': n_tc.get('type','GD')}
        results.append({'id': f'r{rid:03d}', 'trace': r['trace'].zfill(6), 'sequence': r['seq'],
                        'direction': 'Den', 'amount': r['amount'], 'day': day_label,
                        'swift': {'date': sd, 'txnDate': r['txn_date'], 'status': st},
                        'core': entry_c, 'napas': entry_n,
                        'recon_status': rs})
        rid += 1

# ── Print JS output ───────────────────────────────────────────────────────────
print(f'/* Generated from real example files - {len(results)} rows across 3 days */')
print('export const INITIAL_ROWS = [')

def js_val(v):
    if v is None: return 'null'
    if isinstance(v, bool): return 'true' if v else 'false'
    if isinstance(v, int): return str(v)
    if isinstance(v, str): return f"'{v}'"
    if isinstance(v, dict):
        pairs = ', '.join(f"{k}: {js_val(vv)}" for k, vv in v.items())
        return '{' + pairs + '}'
    return repr(v)

def fmt_amount(a):
    s = str(a); parts = []; n = s
    while len(n) > 3: parts.insert(0, n[-3:]); n = n[:-3]
    parts.insert(0, n)
    return '_'.join(parts)

day_dir_prev = None
for r in results:
    key = (r['day'], r['direction'])
    if key != day_dir_prev:
        dir_label = 'Di' if r['direction'] == 'Di' else 'Den'
        print(f"  /* -- {r['day']} {dir_label} */")
        day_dir_prev = key

    sw = js_val(r['swift']) if r['swift'] else 'null'
    co = js_val(r['core'])  if r['core']  else 'null'
    na = js_val(r['napas']) if r['napas'] else 'null'
    amt = fmt_amount(r['amount'])
    seq_js = f"'{r['sequence']}'" if r['sequence'] else 'null'
    dir_js = 'Di' if r['direction'] == 'Di' else 'Den'

    print(f"  {{ id: '{r['id']}', trace: '{r['trace']}', sequence: {seq_js}, direction: '{dir_js}', amount: {amt}, day: '{r['day']}',")
    print(f"    swift: {sw}, core: {co}, napas: {na},")
    print(f"    recon_status: '{r['recon_status']}', resolved_by: null, resolved_at: null, note: null }},")

print(']')
print(f'# Total: {len(results)} rows', file=sys.stderr)
