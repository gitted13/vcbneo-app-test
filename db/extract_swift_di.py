"""
extract_swift_di.py – Đọc "Swift report đi.xlsx" → insert vào swift_raw (direction='Di').

Dùng:
    python extract_swift_di.py <file.xlsx> [--year 2026]

Tự động:
  - Phát hiện hàng header bằng keyword (SEQ, TRACE, HOSTDATE)
  - Phát hiện vị trí cột theo tên header (không hardcode index)
  - Xử lý tất cả sheet trong file
  - Lấy period từ tên sheet DD.MM nếu có, nếu không lấy từ host_date
"""
import sys, io, re
from datetime import date, datetime
import openpyxl
from db_config import get_conn

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Args ──────────────────────────────────────────────────────────────────────
if len(sys.argv) < 2:
    print('Dùng: python extract_swift_di.py <file.xlsx> [--year YYYY]', file=sys.stderr)
    sys.exit(1)

FILE_PATH = sys.argv[1]
YEAR = int(next((a.split('=')[1] if '=' in a else sys.argv[sys.argv.index(a)+1]
                 for a in sys.argv if '--year' in a), date.today().year))

# ── Helpers ───────────────────────────────────────────────────────────────────
def find_col(headers: list[str], *keywords, exclude=()):
    """Tìm index cột theo keyword (case-insensitive), bỏ qua nếu chứa exclude."""
    kws = [k.upper() for k in keywords]
    exs = [e.upper() for e in exclude]
    for i, h in enumerate(headers):
        hu = str(h).upper()
        if any(k in hu for k in kws) and not any(e in hu for e in exs):
            return i
    return None

def find_header_row(ws, keywords=('SEQ', 'TRACE', 'HOSTDATE'), max_row=15):
    """Quét tìm hàng có chứa keyword header."""
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        vals = [str(v).upper() for v in row if v is not None]
        if any(any(kw in v for kw in keywords) for v in vals):
            return row
    return None

def to_int(v):
    if v is None: return None
    try: return int(float(str(v)))
    except: return None

def parse_time_di(raw):
    """'M/D/YYYY H:MM:SS AM/PM' → (date, 'HH:MM') hoặc None"""
    try:
        s = str(raw).strip()
        dp, tp = s.split(' ', 1)
        m, d, y = dp.split('/')
        parts = tp.strip().split(' ')
        hms = parts[0].split(':')
        h, mn = int(hms[0]), int(hms[1])
        if len(parts) > 1:
            if parts[1] == 'PM' and h != 12: h += 12
            elif parts[1] == 'AM' and h == 12: h = 0
        return date(int(y), int(m), int(d)), f'{h:02d}:{mn:02d}'
    except: return None, None

def parse_hostdate(raw):
    """YYYYMMDD int/str hoặc 'YYYY-MM-DD ...' → date"""
    if raw is None: return None
    s = str(raw).strip()
    # YYYYMMDD
    digits = re.sub(r'\D', '', s)[:8]
    if len(digits) == 8:
        try: return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        except: pass
    # ISO format
    for fmt in ('%Y-%m-%d %H:%M:%S', '%Y-%m-%d', '%d/%m/%Y'):
        try: return datetime.strptime(s.split(' ')[0] if ' ' in s else s, fmt.split(' ')[0]).date()
        except: pass
    return None

def sheet_to_period(name: str, year: int):
    """'01.02' → date(year, 2, 1), None nếu không đúng định dạng."""
    try:
        dd, mm = name[:5].split('.')
        return date(year, int(mm), int(dd))
    except: return None

def swift_status(raw) -> str:
    u = str(raw).upper() if raw else ''
    if 'THANH' in u: return 'THANH_CONG'
    if 'TIMEOUT' in u: return 'TIMEOUT'
    return 'THAT_BAI'

# ── Đọc file ──────────────────────────────────────────────────────────────────
print(f'Đọc file: {FILE_PATH}', file=sys.stderr)
wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
rows_out = []

for sh_name in wb.sheetnames:
    ws = wb[sh_name]
    period = sheet_to_period(sh_name, YEAR)

    # Tìm hàng header
    hdr = find_header_row(ws)
    if hdr is None:
        print(f'  Sheet "{sh_name}": không tìm thấy header, bỏ qua.', file=sys.stderr)
        continue

    headers = [str(v) if v is not None else '' for v in hdr]
    hdr_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if list(row) == list(hdr):
            hdr_row_idx = i
            break

    # Xác định cột
    c_time  = find_col(headers, 'THỜI GIAN', 'TIME', 'GIỜ GD')
    c_trace = find_col(headers, 'TRACE NUMBER', 'TRACE NUM', 'TRACE', exclude=('NUMBER OF',))
    c_amt   = find_col(headers, 'SỐ TIỀN', 'AMOUNT', exclude=('PHÍ', 'FEE'))
    c_seq   = find_col(headers, 'SEQ', 'SEQUENCE', exclude=('TRACE',))
    c_hdate = find_col(headers, 'HOSTDATE', 'HOST DATE', 'HOST_DATE')
    c_stat  = find_col(headers, 'PHẢN HỒI', 'KẾT QUẢ', 'STATUS', exclude=('TRẠNG THÁI',))

    missing = [n for n, c in [('TRACE',c_trace),('SỐ TIỀN',c_amt),('HOSTDATE',c_hdate)] if c is None]
    if missing:
        print(f'  Sheet "{sh_name}": thiếu cột {missing}, bỏ qua.', file=sys.stderr)
        continue

    count = 0
    for row in ws.iter_rows(min_row=(hdr_row_idx or 7) + 1, values_only=True):
        if all(v is None for v in row): continue

        txn_date, _ = parse_time_di(row[c_time]) if c_time is not None and row[c_time] else (None, None)
        trace   = str(to_int(row[c_trace])).zfill(6) if to_int(row[c_trace]) else None
        amt     = to_int(row[c_amt]) or 0
        seq     = str(to_int(row[c_seq])) if c_seq is not None and to_int(row[c_seq]) else None
        hdate   = parse_hostdate(row[c_hdate]) if c_hdate is not None else None
        st      = swift_status(row[c_stat] if c_stat is not None else None)

        if not trace or not amt: continue
        host_date = hdate or txn_date
        p = period or host_date
        if not p or not host_date or not txn_date: continue

        rows_out.append((p, 'Di', trace, seq, txn_date, host_date, amt, st))
        count += 1

    print(f'  Sheet "{sh_name}": {count} dòng.', file=sys.stderr)

wb.close()
print(f'Tổng: {len(rows_out)} dòng Swift Đi.', file=sys.stderr)

# ── Insert vào DB ─────────────────────────────────────────────────────────────
if not rows_out:
    print('Không có dữ liệu để insert.', file=sys.stderr)
    sys.exit(0)

conn = get_conn()
cur  = conn.cursor()
sql = """
    MERGE swift_raw AS tgt
    USING (VALUES (?,?,?,?,?,?,?,?)) AS src
        (period, direction, trace, sequence, txn_date, host_date, amount, status)
    ON  tgt.period = src.period AND tgt.direction = src.direction
    AND tgt.trace  = src.trace  AND tgt.amount    = src.amount
    WHEN MATCHED THEN UPDATE SET
        sequence=src.sequence, txn_date=src.txn_date,
        host_date=src.host_date, status=src.status
    WHEN NOT MATCHED THEN INSERT
        (period, direction, trace, sequence, txn_date, host_date, amount, status)
    VALUES
        (src.period, src.direction, src.trace, src.sequence,
         src.txn_date, src.host_date, src.amount, src.status);
"""
cur.executemany(sql, rows_out)
conn.commit()
conn.close()
print('Done.', file=sys.stderr)
