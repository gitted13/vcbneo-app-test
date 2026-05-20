"""
extract_napas_ktc.py – Đọc file "Napas đi KTC *.xlsx" → insert vào napas_raw (failed=1).

Dùng:
    python extract_napas_ktc.py <file.xlsx> [--year 2026]

KTC = Không Thành Công. failed=1, napas_type='GD'.
Cấu trúc tương tự NAPAS Đi nhưng cột trace ở vị trí khác (col 3 thay vì col 2).
"""
import sys, io
from datetime import date
import openpyxl
from db_config import get_conn

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

if len(sys.argv) < 2:
    print('Dùng: python extract_napas_ktc.py <file.xlsx> [--year YYYY]', file=sys.stderr)
    sys.exit(1)

FILE_PATH = sys.argv[1]
YEAR = int(next((a.split('=')[1] if '=' in a else sys.argv[sys.argv.index(a)+1]
                 for a in sys.argv if '--year' in a), date.today().year))

def find_col(headers, *keywords, exclude=()):
    kws = [k.upper() for k in keywords]
    exs = [e.upper() for e in exclude]
    for i, h in enumerate(headers):
        hu = str(h).upper()
        if any(k in hu for k in kws) and not any(e in hu for e in exs):
            return i
    return None

def to_int(v):
    if v is None: return None
    try: return int(float(str(v)))
    except: return None

def napas_time(raw):
    try:
        s = str(int(raw)).zfill(6)
        return f'{s[0:2]}:{s[2:4]}'
    except: return None

def napas_date(raw, year: int, p_mmdd: str):
    try:
        s = str(int(raw)).zfill(4)
        mm, dd = int(s[0:2]), int(s[2:4])
        d = date(year, mm, dd)
        return d, 'GD' if s == p_mmdd else 'QT'
    except: return None, 'GD'

def sheet_to_period(name: str, year: int):
    try:
        dd, mm = name[:5].split('.')
        return date(year, int(mm), int(dd))
    except: return None

def sheet_mmdd(name: str):
    try:
        dd, mm = name[:5].split('.')
        return f'{mm.zfill(2)}{dd.zfill(2)}'
    except: return '0000'

print(f'Đọc file: {FILE_PATH}', file=sys.stderr)
wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
rows_out = []

for sh_name in wb.sheetnames:
    ws = wb[sh_name]
    period = sheet_to_period(sh_name, YEAR)
    p_mmdd = sheet_mmdd(sh_name)

    all_rows = list(ws.iter_rows(min_row=1, max_row=3, values_only=True))
    if not all_rows: continue
    headers = [str(v) if v is not None else '' for v in all_rows[0]]

    c_amt   = find_col(headers, 'SỐ TIỀN', 'TIỀN', 'AMOUNT', exclude=('PHÍ', 'FEE'))
    c_trace = find_col(headers, 'TRACE', 'SỐ TRACE')
    c_time  = find_col(headers, 'GIỜ', 'TIME', 'HHMMSS')
    c_date  = find_col(headers, 'NGÀY GD', 'NGÀY GIAO DỊCH', 'MMDD', 'NGÀY')

    if c_trace is None or c_amt is None:
        # fallback KTC: 0=STT, 2=Số tiền, 3=Số trace, 4=Giờ, 5=Ngày GD
        c_amt, c_trace, c_time, c_date = 2, 3, 4, 5
        print(f'  Sheet "{sh_name}": dùng vị trí cột mặc định.', file=sys.stderr)

    count = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row): continue
        trace = str(to_int(row[c_trace])).zfill(6) if to_int(row[c_trace]) else None
        amt   = to_int(row[c_amt]) or 0
        if not trace or not amt: continue

        raw_date = row[c_date] if c_date is not None and len(row) > c_date else None
        txn_date, _ = napas_date(raw_date, YEAR, p_mmdd) if raw_date else (period, 'GD')
        ntime = napas_time(row[c_time]) if c_time is not None and len(row) > c_time and row[c_time] else None
        p = period or txn_date
        if not p or not txn_date: continue

        rows_out.append((p, 'Di', trace, txn_date, ntime, amt, 1, 'GD'))
        count += 1

    print(f'  Sheet "{sh_name}": {count} dòng KTC.', file=sys.stderr)

wb.close()
print(f'Tổng: {len(rows_out)} dòng NAPAS KTC.', file=sys.stderr)

if not rows_out:
    print('Không có dữ liệu để insert.', file=sys.stderr)
    sys.exit(0)

conn = get_conn()
cur  = conn.cursor()
sql = """
    MERGE napas_raw AS tgt
    USING (VALUES (?,?,?,?,?,?,?,?)) AS src
        (period, direction, trace, txn_date, txn_time, amount, failed, napas_type)
    ON  tgt.period = src.period AND tgt.direction = src.direction
    AND tgt.trace  = src.trace  AND tgt.amount    = src.amount
    WHEN MATCHED THEN UPDATE SET
        txn_date=src.txn_date, txn_time=src.txn_time,
        failed=src.failed, napas_type=src.napas_type
    WHEN NOT MATCHED THEN INSERT
        (period, direction, trace, txn_date, txn_time, amount, failed, napas_type)
    VALUES
        (src.period, src.direction, src.trace, src.txn_date, src.txn_time,
         src.amount, src.failed, src.napas_type);
"""
cur.executemany(sql, rows_out)
conn.commit()
conn.close()
print('Done.', file=sys.stderr)
