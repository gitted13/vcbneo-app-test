"""
extract_swift_den.py – Đọc "Swift report đến.xlsx" → insert vào swift_raw (direction='Den').

Dùng:
    python extract_swift_den.py <file.xlsx> [--year 2026]
"""
import sys, io, re
from datetime import date, datetime
import openpyxl
from db_config import get_conn

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

if len(sys.argv) < 2:
    print('Dùng: python extract_swift_den.py <file.xlsx> [--year YYYY]', file=sys.stderr)
    sys.exit(1)

FILE_PATH = sys.argv[1]
YEAR = int(next((a.split('=')[1] if '=' in a else sys.argv[sys.argv.index(a)+1]
                 for a in sys.argv if '--year' in a), date.today().year))

# ── Helpers ───────────────────────────────────────────────────────────────────
def find_col(headers, *keywords, exclude=()):
    kws = [k.upper() for k in keywords]
    exs = [e.upper() for e in exclude]
    for i, h in enumerate(headers):
        hu = str(h).upper()
        if any(k in hu for k in kws) and not any(e in hu for e in exs):
            return i
    return None

def find_header_row(ws, keywords=('TRACE', 'SỐ TIỀN', 'HOSTDATE'), max_row=15):
    for row in ws.iter_rows(min_row=1, max_row=max_row, values_only=True):
        vals = [str(v).upper() for v in row if v is not None]
        if any(any(kw in v for kw in keywords) for v in vals):
            return row
    return None

def to_int(v):
    if v is None: return None
    try: return int(float(str(v)))
    except: return None

def parse_time_den(raw):
    """'YYYY-MM-DD HH:MM:SS' → (date, 'HH:MM')"""
    try:
        s = str(raw).strip()
        dp, tp = s.split(' ')
        y, mo, d = dp.split('-')
        h, mn = tp.split(':')[0], tp.split(':')[1]
        return date(int(y), int(mo), int(d)), f'{int(h):02d}:{int(mn):02d}'
    except: return None, None

def parse_hostdate(raw):
    if raw is None: return None
    s = str(raw).strip()
    # ISO 'YYYY-MM-DD ...'
    try:
        p = s.split(' ')[0]
        parts = p.split('-')
        if len(parts) == 3:
            return date(int(parts[0]), int(parts[1]), int(parts[2][:2]))
    except: pass
    # YYYYMMDD digits
    digits = re.sub(r'\D', '', s)[:8]
    if len(digits) == 8:
        try: return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        except: pass
    return None

def sheet_to_period(name: str, year: int):
    try:
        dd, mm = name[:5].split('.')
        return date(year, int(mm), int(dd))
    except: return None

def swift_status(raw) -> str:
    u = str(raw).upper() if raw else ''
    if 'THANH' in u: return 'THANH_CONG'
    if 'TIMEOUT' in u: return 'TIMEOUT'
    return 'THAT_BAI'

# ── Đọc file ──────────────────────────────────────────────────────────────────
print(f'Đọc file: {FILE_PATH}', file=sys.stderr)
wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
rows_out = []

for sh_name in wb.sheetnames:
    ws = wb[sh_name]
    period = sheet_to_period(sh_name, YEAR)

    hdr = find_header_row(ws)
    if hdr is None:
        print(f'  Sheet "{sh_name}": không tìm thấy header, bỏ qua.', file=sys.stderr)
        continue

    headers = [str(v) if v is not None else '' for v in hdr]
    hdr_row_idx = None
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=20, values_only=True), 1):
        if list(row) == list(hdr):
            hdr_row_idx = i
            break

    c_time  = find_col(headers, 'THỜI GIAN', 'TIME', 'GIỜ GD')
    c_amt   = find_col(headers, 'SỐ TIỀN', 'AMOUNT', exclude=('PHÍ', 'FEE'))
    c_hdate = find_col(headers, 'HOSTDATE', 'HOST DATE', 'HOST_DATE')
    c_trace = find_col(headers, 'TRACE', exclude=('NUMBER OF',))
    c_seq   = find_col(headers, 'SEQ', 'SEQUENCE', exclude=('TRACE',))
    c_stat  = find_col(headers, 'PHẢN HỒI', 'KẾT QUẢ', 'STATUS', exclude=('TRẠNG THÁI',))

    missing = [n for n, c in [('TRACE',c_trace),('SỐ TIỀN',c_amt)] if c is None]
    if missing:
        print(f'  Sheet "{sh_name}": thiếu cột {missing}, bỏ qua.', file=sys.stderr)
        continue

    count = 0
    for row in ws.iter_rows(min_row=(hdr_row_idx or 7) + 1, values_only=True):
        if all(v is None for v in row): continue

        txn_date, _ = parse_time_den(row[c_time]) if c_time is not None and row[c_time] else (None, None)
        amt   = to_int(row[c_amt]) or 0
        hdate = parse_hostdate(row[c_hdate]) if c_hdate is not None and row[c_hdate] else None
        trace = str(to_int(row[c_trace])).zfill(6) if to_int(row[c_trace]) else None
        seq   = str(to_int(row[c_seq])) if c_seq is not None and to_int(row[c_seq]) else None
        st    = swift_status(row[c_stat] if c_stat is not None else None)

        if not trace or not amt: continue
        host_date = hdate or txn_date
        p = period or host_date
        if not p or not host_date or not txn_date: continue

        rows_out.append((p, 'Den', trace, seq, txn_date, host_date, amt, st))
        count += 1

    print(f'  Sheet "{sh_name}": {count} dòng.', file=sys.stderr)

wb.close()
print(f'Tổng: {len(rows_out)} dòng Swift Đến.', file=sys.stderr)

if not rows_out:
    print('Không có dữ liệu để insert.', file=sys.stderr)
    sys.exit(0)

conn = get_conn()
cur  = conn.cursor()
sql = """
    MERGE swift_raw AS tgt
    USING (VALUES (?,?,?,?,?,?,?,?)) AS src
        (period, direction, trace, sequence, txn_date, host_date, amount, status)
    ON  tgt.period = src.period AND tgt.direction = src.direction
    AND tgt.trace  = src.trace  AND tgt.amount    = src.amount
    WHEN MATCHED THEN UPDATE SET
        sequence=src.sequence, txn_date=src.txn_date,
        host_date=src.host_date, status=src.status
    WHEN NOT MATCHED THEN INSERT
        (period, direction, trace, sequence, txn_date, host_date, amount, status)
    VALUES
        (src.period, src.direction, src.trace, src.sequence,
         src.txn_date, src.host_date, src.amount, src.status);
"""
cur.executemany(sql, rows_out)
conn.commit()
conn.close()
print('Done.', file=sys.stderr)
