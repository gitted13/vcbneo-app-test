"""
extract_to_db.py  –  Đọc file Excel → insert vào bảng raw MSSQL.
Triggers tự động populate swift_core_recon, napas_core_recon, master_recon.

Dùng:
    python extract_to_db.py                          # quét tất cả file trong EXAMPLE_DIR
    python extract_to_db.py swift_di=/path/file.xlsx # chỉ định file cụ thể
    python extract_to_db.py swift_di napas_di        # chỉ chạy 2 nguồn có sẵn

Tham số dòng lệnh:
    slot=path  → dùng file ở path cho slot đó
    slot       → bỏ qua slot đó (chạy các slot còn lại trong EXAMPLE_DIR)

Slots hợp lệ: swift_di, swift_den, napas_di, napas_den, core, napas_ktc

Yêu cầu:
    pip install openpyxl pyodbc
"""
import openpyxl, re, glob, os, sys, io
from datetime import date
import pyodbc

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

# ── Cấu hình ──────────────────────────────────────────────────────────────────
CONNECTION_STRING = (
    "DRIVER={ODBC Driver 17 for SQL Server};"
    "SERVER=localhost;"
    "DATABASE=vcbneo;"
    "Trusted_Connection=yes;"
)

EXAMPLE_DIR = r'd:\FoxAI\FoxAI\*\vcbneo-app\example\*'
YEAR = 2026

# ── Parse tham số dòng lệnh ───────────────────────────────────────────────────
# Dạng: slot=path  hoặc  slot (chỉ blacklist slot đó)
forced  = {}   # slot → path do người dùng chỉ định
skipped = set()  # slot bị bỏ qua

for arg in sys.argv[1:]:
    if '=' in arg:
        k, v = arg.split('=', 1)
        forced[k.strip()] = v.strip()
    else:
        skipped.add(arg.strip())

# ── Tìm file Excel trong EXAMPLE_DIR ─────────────────────────────────────────
NAME_MAP = {
    b'Swift report \xc4\x91i.xlsx':             'swift_di',
    b'Swift report \xc4\x91\xe1\xba\xbfn.xlsx': 'swift_den',
    b'Napas \xc4\x91i.xlsx':                    'napas_di',
    b'Napas \xc4\x91\xe1\xba\xbfn.xlsx':        'napas_den',
    b'Core.xlsx':                                'core',
    b'Napas \xc4\x91i KTC 20260201.XLSX':       'napas_ktc',
}

fmap = dict(forced)  # bắt đầu bằng file do người dùng chỉ định
for f in glob.glob(EXAMPLE_DIR):
    enc = os.path.basename(f).encode('utf-8')
    slot = NAME_MAP.get(enc)
    if slot and slot not in fmap and slot not in skipped:
        fmap[slot] = f

present = [s for s in ('swift_di', 'swift_den', 'napas_di', 'napas_den', 'core', 'napas_ktc') if s in fmap]
print(f'File tìm thấy: {present}', file=sys.stderr)
if not present:
    print('Không tìm thấy file nào để xử lý.', file=sys.stderr)
    sys.exit(0)

# ── Patterns Core GL ──────────────────────────────────────────────────────────
CRED_PAT = re.compile(r'06800-(\d+)-(\d+)\s+TRANSFER\s+CREDMBNEO\.(\d+)\.(\d+)')
DEB_PAT  = re.compile(r'06800-(\d+)-(\d+)\s+TRANSFER\s+')

# ── Helpers ───────────────────────────────────────────────────────────────────
def to_int(v):
    if v is None: return None
    try: return int(float(str(v)))
    except: return None

def to_str(v):
    if v is None: return ''
    iv = to_int(v)
    return str(iv) if iv is not None else str(v).strip()

def parse_date_ddmmyyyy(s):
    if not s: return None
    try:
        d, m, y = s.split('/')
        return date(int(y), int(m), int(d))
    except: return None

def dt8_to_date(raw):
    s = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
    if len(s) == 8:
        return date(int(s[0:4]), int(s[4:6]), int(s[6:8]))
    return None

def dt_iso_to_date(s):
    p = str(s).strip().split(' ')[0]
    if '-' in p:
        parts = p.split('-')
        return date(int(parts[0]), int(parts[1]), int(parts[2][:2]))
    return None

def parse_swift_di_time(raw):
    try:
        s = str(raw).strip()
        date_p, time_p = s.split(' ', 1)
        m, d, y = date_p.split('/')
        tp = time_p.strip().split(' ')
        hms = tp[0].split(':')
        h, mn = int(hms[0]), int(hms[1])
        if len(tp) > 1:
            if tp[1] == 'PM' and h != 12: h += 12
            elif tp[1] == 'AM' and h == 12: h = 0
        return date(int(y), int(m), int(d)), f'{h:02d}:{mn:02d}'
    except: return None, None

def parse_swift_den_time(raw):
    try:
        s = str(raw).strip()
        dp, tp = s.split(' ')
        y, mo, d = dp.split('-')
        h, mn = tp.split(':')[0], tp.split(':')[1]
        return date(int(y), int(mo), int(d)), f'{int(h):02d}:{int(mn):02d}'
    except: return None, None

def napas_time_fmt(raw):
    try:
        s = str(int(raw)).zfill(6)
        return f'{s[0:2]}:{s[2:4]}'
    except: return None

def napas_date_type(raw, p_mmdd):
    try:
        s = str(int(raw)).zfill(4)
        mm, dd = int(s[0:2]), int(s[2:4])
        d = date(YEAR, mm, dd)
        return d, 'GD' if s[0:4] == p_mmdd else 'QT'
    except: return None, 'GD'

def sheet_to_date(sh):
    try:
        dd, mm = sh.split('.')
        return date(YEAR, int(mm[:2]), int(dd))
    except: return None

# ── Kết nối DB ────────────────────────────────────────────────────────────────
print('Connecting to SQL Server...', file=sys.stderr)
conn = pyodbc.connect(CONNECTION_STRING)
cur  = conn.cursor()

# ── Insert helpers (MERGE = upsert) ───────────────────────────────────────────
def insert_swift(rows):
    if not rows: return
    sql = """
        MERGE swift_raw AS tgt
        USING (VALUES (?,?,?,?,?,?,?,?)) AS src
            (period, direction, trace, sequence, txn_date, host_date, amount, status)
        ON  tgt.period    = src.period    AND tgt.direction = src.direction
        AND tgt.trace     = src.trace     AND tgt.amount    = src.amount
        WHEN NOT MATCHED THEN INSERT
            (period, direction, trace, sequence, txn_date, host_date, amount, status)
        VALUES
            (src.period, src.direction, src.trace, src.sequence,
             src.txn_date, src.host_date, src.amount, src.status);
    """
    cur.executemany(sql, rows)

def insert_napas(rows):
    if not rows: return
    sql = """
        MERGE napas_raw AS tgt
        USING (VALUES (?,?,?,?,?,?,?,?)) AS src
            (period, direction, trace, txn_date, txn_time, amount, failed, napas_type)
        ON  tgt.period    = src.period    AND tgt.direction = src.direction
        AND tgt.trace     = src.trace     AND tgt.amount    = src.amount
        WHEN NOT MATCHED THEN INSERT
            (period, direction, trace, txn_date, txn_time, amount, failed, napas_type)
        VALUES
            (src.period, src.direction, src.trace, src.txn_date, src.txn_time,
             src.amount, src.failed, src.napas_type);
    """
    cur.executemany(sql, rows)

def insert_core(rows):
    if not rows: return
    sql = """
        MERGE core_raw AS tgt
        USING (VALUES (?,?,?,?,?,?,?,?)) AS src
            (period, direction, trace, sequence, txn_date, amount, entry_type, description)
        ON  tgt.period    = src.period    AND tgt.direction = src.direction
        AND tgt.sequence  = src.sequence  AND tgt.amount    = src.amount
        WHEN NOT MATCHED THEN INSERT
            (period, direction, trace, sequence, txn_date, amount, entry_type, description)
        VALUES
            (src.period, src.direction, src.trace, src.sequence,
             src.txn_date, src.amount, src.entry_type, src.description);
    """
    cur.executemany(sql, rows)

# ── Xử lý từng nguồn (chỉ khi file tồn tại) ─────────────────────────────────

# NAPAS Đi
napas_rows = []
if 'napas_di' in fmap:
    print('Reading NAPAS Đi...', file=sys.stderr)
    wb = openpyxl.load_workbook(fmap['napas_di'], read_only=True, data_only=True)
    for sh in wb.sheetnames:
        period = sheet_to_date(sh[:5])
        p_mmdd = sh[:4]
        for row in wb[sh].iter_rows(min_row=2, values_only=True):
            trace = to_str(row[2]); amt = to_int(row[1]) or 0
            if not trace or not amt: continue
            txn_date, typ = napas_date_type(row[4], p_mmdd) if row[4] else (period, 'GD')
            ntime = napas_time_fmt(row[3]) if row[3] else None
            napas_rows.append((period, 'Di', trace.zfill(6), txn_date, ntime, amt, 0, typ))
    wb.close()

# NAPAS Đi KTC
if 'napas_ktc' in fmap:
    print('Reading NAPAS Đi KTC...', file=sys.stderr)
    wb = openpyxl.load_workbook(fmap['napas_ktc'], read_only=True, data_only=True)
    for sh in wb.sheetnames:
        period = sheet_to_date(sh[:5])
        p_mmdd = sh[:4]
        for row in wb[sh].iter_rows(min_row=2, values_only=True):
            trace = to_str(row[3]); amt = to_int(row[2]) or 0
            if not trace or not amt: continue
            txn_date, _ = napas_date_type(row[5], p_mmdd) if row[5] else (period, 'GD')
            ntime = napas_time_fmt(row[4]) if row[4] else None
            napas_rows.append((period, 'Di', trace.zfill(6), txn_date, ntime, amt, 1, 'GD'))
    wb.close()

# NAPAS Đến
if 'napas_den' in fmap:
    print('Reading NAPAS Đến...', file=sys.stderr)
    wb = openpyxl.load_workbook(fmap['napas_den'], read_only=True, data_only=True)
    for sh in wb.sheetnames:
        period = sheet_to_date(sh[:5])
        p_mmdd = sh[:4]
        for row in wb[sh].iter_rows(min_row=2, values_only=True):
            trace = to_str(row[3]); amt = to_int(row[2]) or 0
            if not trace or not amt: continue
            txn_date, typ = napas_date_type(row[5], p_mmdd) if row[5] else (period, 'GD')
            ntime = napas_time_fmt(row[4]) if row[4] else None
            napas_rows.append((period, 'Den', trace.zfill(6), txn_date, ntime, amt, 0, typ))
    wb.close()

insert_napas(napas_rows)
print(f'  → {len(napas_rows)} NAPAS rows inserted/skipped', file=sys.stderr)

# Core GL
core_rows = []
if 'core' in fmap:
    print('Reading Core GL...', file=sys.stderr)
    wb = openpyxl.load_workbook(fmap['core'], read_only=True, data_only=True)
    for sh in wb.sheetnames:
        for row in wb[sh].iter_rows(min_row=6, values_only=True):
            if row[1] is None or row[7] is None: continue
            debit  = float(row[4]) if row[4] else 0
            credit = float(row[5]) if row[5] else 0
            desc   = str(row[7])
            txn_date = dt8_to_date(row[1])
            if not txn_date: continue
            if credit > 0:
                m = CRED_PAT.search(desc)
                if m:
                    seq = m.group(2)
                    core_rows.append((txn_date, 'Di', m.group(4).zfill(6), seq, txn_date, int(credit), 'Ghi co', desc[:500]))
            elif debit > 0:
                m = DEB_PAT.search(desc)
                if m:
                    seq = m.group(2)
                    core_rows.append((txn_date, 'Den', None, seq, txn_date, int(debit), 'Ghi no', desc[:500]))
    wb.close()

insert_core(core_rows)
print(f'  → {len(core_rows)} Core rows inserted/skipped', file=sys.stderr)

# Swift Đi
swift_rows = []
if 'swift_di' in fmap:
    print('Reading Swift Đi...', file=sys.stderr)
    wb = openpyxl.load_workbook(fmap['swift_di'], read_only=True, data_only=True)
    for sh in wb.sheetnames:
        period = sheet_to_date(sh[:5])
        if not period: continue
        for row in wb[sh].iter_rows(min_row=8, values_only=True):
            txn_date, _ = parse_swift_di_time(row[0]) if row[0] else (None, None)
            trace = to_str(row[6]);  amt = to_int(row[7]) or 0
            seq   = to_str(row[12]); hd  = to_str(row[14])
            if not trace or not amt: continue
            host_date = dt8_to_date(hd) if hd and len(hd) == 8 else txn_date
            st_raw = str(row[16]) if row[16] else ''
            st = 'THANH_CONG' if 'THANH' in st_raw else ('TIMEOUT' if 'TIMEOUT' in st_raw else 'THAT_BAI')
            if not txn_date or not host_date: continue
            swift_rows.append((period, 'Di', trace.zfill(6), seq or None,
                               txn_date, host_date, amt, st))
    wb.close()

# Swift Đến
if 'swift_den' in fmap:
    print('Reading Swift Đến...', file=sys.stderr)
    wb = openpyxl.load_workbook(fmap['swift_den'], read_only=True, data_only=True)
    for sh in wb.sheetnames:
        period = sheet_to_date(sh[:5])
        if not period: continue
        for row in wb[sh].iter_rows(min_row=8, values_only=True):
            txn_date, _ = parse_swift_den_time(row[1]) if row[1] else (None, None)
            amt   = to_int(row[8]) or 0
            hd    = str(row[10]) if row[10] else ''
            trace = to_str(row[11]); seq = to_str(row[13])
            if not trace or not amt: continue
            host_date = dt_iso_to_date(hd) if hd else txn_date
            st_raw = str(row[16]) if row[16] else ''
            st = 'THANH_CONG' if 'THANH' in st_raw else ('TIMEOUT' if 'TIMEOUT' in st_raw else 'THAT_BAI')
            if not txn_date or not host_date: continue
            swift_rows.append((period, 'Den', trace.zfill(6), seq or None,
                               txn_date, host_date, amt, st))
    wb.close()

insert_swift(swift_rows)
print(f'  → {len(swift_rows)} Swift rows inserted/skipped', file=sys.stderr)

# ── Commit ────────────────────────────────────────────────────────────────────
conn.commit()
conn.close()
print('Done.', file=sys.stderr)
