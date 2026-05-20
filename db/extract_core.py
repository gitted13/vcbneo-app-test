"""
extract_core.py – Đọc "Core.xlsx" → insert vào core_raw.

Dùng:
    python extract_core.py <file.xlsx> [--year 2026]

Logic:
  - Tìm hàng header có "DIỄN GIẢI" hoặc "NGÀY GIAO DỊCH"
  - Dòng Ghi có (CREDIT > 0): parse CRED_PAT → direction='Di', entry='Ghi co'
  - Dòng Ghi nợ (DEBIT > 0):  parse DEB_PAT  → direction='Den', entry='Ghi no'
  - Khớp theo sequence + amount với swift_raw (trigger xử lý tự động)
"""
import sys, io, re
from datetime import date
import openpyxl
from db_config import get_conn

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

if len(sys.argv) < 2:
    print('Dùng: python extract_core.py <file.xlsx> [--year YYYY]', file=sys.stderr)
    sys.exit(1)

FILE_PATH = sys.argv[1]
YEAR = int(next((a.split('=')[1] if '=' in a else sys.argv[sys.argv.index(a)+1]
                 for a in sys.argv if '--year' in a), date.today().year))

CRED_PAT = re.compile(r'06800-(\d+)-(\d+)\s+TRANSFER\s+CREDMBNEO\.(\d+)\.(\d+)', re.IGNORECASE)
DEB_PAT  = re.compile(r'06800-(\d+)-(\d+)\s+TRANSFER\s+',                         re.IGNORECASE)

# ── Helpers ───────────────────────────────────────────────────────────────────
def find_col(headers, *keywords, exclude=()):
    kws = [k.upper() for k in keywords]
    exs = [e.upper() for e in exclude]
    for i, h in enumerate(headers):
        hu = str(h).upper()
        if any(k in hu for k in kws) and not any(e in hu for e in exs):
            return i
    return None

def find_header_row(ws, max_row=20):
    """Tìm hàng có DIỄN GIẢI hoặc NGÀY GIAO DỊCH."""
    keywords = ('DIỄN GIẢI', 'DIEN GIAI', 'NGÀY GIAO DỊCH', 'NGAY GIAO DICH')
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=max_row, values_only=True), 1):
        vals = [str(v).upper() for v in row if v is not None]
        if any(any(kw in v for kw in keywords) for v in vals):
            return i, list(row)
    return None, None

def to_float(v):
    if v is None: return 0.0
    try: return float(v)
    except: return 0.0

def parse_core_date(raw):
    """YYYYMMDD int → date"""
    if raw is None: return None
    try:
        s = str(int(raw)) if isinstance(raw, (int, float)) else str(raw).strip()
        if len(s) == 8:
            return date(int(s[0:4]), int(s[4:6]), int(s[6:8]))
    except: pass
    return None

# ── Đọc file ──────────────────────────────────────────────────────────────────
print(f'Đọc file: {FILE_PATH}', file=sys.stderr)
wb = openpyxl.load_workbook(FILE_PATH, read_only=True, data_only=True)
rows_out = []

for sh_name in wb.sheetnames:
    ws = wb[sh_name]
    hdr_row, hdr = find_header_row(ws)
    if hdr_row is None:
        print(f'  Sheet "{sh_name}": không tìm thấy header, bỏ qua.', file=sys.stderr)
        continue

    headers = [str(v) if v is not None else '' for v in hdr]
    c_date  = find_col(headers, 'NGÀY GIAO DỊCH', 'NGÀY GD', 'NGAY GIAO DICH', 'DATE')
    c_debit = find_col(headers, 'GHI NỢ', 'DEBIT', 'NỢ', exclude=('CÓ',))
    c_credit= find_col(headers, 'GHI CÓ', 'CREDIT', 'CÓ', exclude=('NỢ',))
    c_desc  = find_col(headers, 'DIỄN GIẢI', 'DIEN GIAI', 'MÔ TẢ', 'DESCRIPTION')

    if c_date is None or c_desc is None:
        # fallback vị trí mặc định: col1=ngày, col4=debit, col5=credit, col7=diễn giải
        c_date, c_debit, c_credit, c_desc = 1, 4, 5, 7
        print(f'  Sheet "{sh_name}": dùng vị trí cột mặc định.', file=sys.stderr)

    count = 0
    for row in ws.iter_rows(min_row=hdr_row + 1, values_only=True):
        if not row or all(v is None for v in row): continue
        txn_date = parse_core_date(row[c_date] if len(row) > c_date else None)
        if not txn_date: continue

        desc   = str(row[c_desc]) if len(row) > c_desc and row[c_desc] else ''
        debit  = to_float(row[c_debit]  if len(row) > c_debit  else None)
        credit = to_float(row[c_credit] if len(row) > c_credit else None)

        if credit > 0:
            m = CRED_PAT.search(desc)
            if not m: continue
            seq   = m.group(2)
            trace = m.group(4).zfill(6)
            rows_out.append((txn_date, 'Di', trace, seq, txn_date, int(credit), 'Ghi co', desc[:500]))
            count += 1
        elif debit > 0:
            m = DEB_PAT.search(desc)
            if not m: continue
            seq = m.group(2)
            rows_out.append((txn_date, 'Den', None, seq, txn_date, int(debit), 'Ghi no', desc[:500]))
            count += 1

    print(f'  Sheet "{sh_name}": {count} dòng.', file=sys.stderr)

wb.close()
print(f'Tổng: {len(rows_out)} dòng Core GL.', file=sys.stderr)

if not rows_out:
    print('Không có dữ liệu để insert.', file=sys.stderr)
    sys.exit(0)

conn = get_conn()
cur  = conn.cursor()
sql = """
    MERGE core_raw AS tgt
    USING (VALUES (?,?,?,?,?,?,?,?)) AS src
        (period, direction, trace, sequence, txn_date, amount, entry_type, description)
    ON  tgt.period   = src.period   AND tgt.direction = src.direction
    AND tgt.sequence = src.sequence AND tgt.amount    = src.amount
    WHEN MATCHED THEN UPDATE SET
        trace=src.trace, txn_date=src.txn_date,
        entry_type=src.entry_type, description=src.description
    WHEN NOT MATCHED THEN INSERT
        (period, direction, trace, sequence, txn_date, amount, entry_type, description)
    VALUES
        (src.period, src.direction, src.trace, src.sequence,
         src.txn_date, src.amount, src.entry_type, src.description);
"""
cur.executemany(sql, rows_out)
conn.commit()
conn.close()
print('Done.', file=sys.stderr)
